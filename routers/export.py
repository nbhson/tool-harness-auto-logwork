"""Router export — xuất dữ liệu ra Excel file."""

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from database.db import get_db
from database.models import WorkLog

router = APIRouter(prefix="/api/export", tags=["export"])


# ─── Styling constants ────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="Calibri", size=10)
CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SOURCE_COLORS = {
    "jira": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "bitbucket": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "git": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "manual": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
}


def _build_workbook(source_filter: Optional[str]) -> io.BytesIO:
    """Tạo Excel workbook từ dữ liệu work logs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Logs"

    # ─── Headers ──────────────────────────────────────
    headers = [
        "ID",
        "Source",
        "Activity Type",
        "Title",
        "Description",
        "Project",
        "URL",
        "Activity Date",
        "Time (hours)",
        "Time (minutes)",
        "External ID",
        "Created At",
    ]

    # Column widths
    col_widths = [6, 12, 16, 40, 50, 18, 50, 20, 12, 14, 24, 20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── Data & Summary (1 session duy nhất) ─────────
    db = get_db()
    try:
        session: Session = next(db)

        # Build query (có hoặc không filter)
        q = session.query(WorkLog).order_by(WorkLog.activity_timestamp.desc())
        if source_filter:
            q = q.filter(WorkLog.source == source_filter)

        # Dùng cùng query cho cả data rows và summary
        all_logs = q.all()
        total_time = sum(log.time_spent_minutes or 0 for log in all_logs)
        total = len(all_logs)

        # ─── Data rows ────────────────────────────────
        for row_idx, log in enumerate(all_logs, 2):
            source_fill = SOURCE_COLORS.get(log.source)

            values = [
                log.id,
                log.source,
                log.activity_type,
                log.title,
                log.description or "",
                log.project or "",
                log.url or "",
                log.activity_timestamp.strftime("%Y-%m-%d %H:%M") if log.activity_timestamp else "",
                round(log.time_spent_minutes / 60, 2) if log.time_spent_minutes else 0,
                log.time_spent_minutes or 0,
                log.external_id or "",
                log.created_at.strftime("%Y-%m-%d %H:%M") if log.created_at else "",
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                if source_fill:
                    cell.fill = source_fill

        # ─── Freeze panes ──────────────────────────────
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ─── Summary sheet (dùng cùng source_filter) ──
        ws2 = wb.create_sheet("Summary")
        summary_headers = ["Metric", "Value"]
        for col_idx, h in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 15

        # Đếm theo source (dựa trên dữ liệu đã filter)
        counts = {}
        for log in all_logs:
            counts[log.source] = counts.get(log.source, 0) + 1

        summary_data = [
            ("Total Logs", total),
            ("Total Time (hours)", round(total_time / 60, 1)),
            ("Total Time (minutes)", total_time),
            ("Jira Activities", counts.get("jira", 0)),
            ("Bitbucket Activities", counts.get("bitbucket", 0)),
            ("GitHub Activities", counts.get("github", 0)),
            ("Git Commits", counts.get("git", 0)),
            ("Manual Entries", counts.get("manual", 0)),
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws2.cell(row=row_idx, column=1, value=metric).font = CELL_FONT
            ws2.cell(row=row_idx, column=2, value=value).font = CELL_FONT
            ws2.cell(row=row_idx, column=1).border = THIN_BORDER
            ws2.cell(row=row_idx, column=2).border = THIN_BORDER

    finally:
        db.close()

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/excel")
def export_excel(
    source: Optional[str] = Query(None, description="Filter by source"),
):
    """Export work logs ra file Excel (.xlsx).

    Trả về file Excel với 2 sheets:
    - Work Logs: chi tiết từng entry
    - Summary: thống kê tổng quan
    """
    output = _build_workbook(source)

    filename = f"worklog_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/excel/{source}")
def export_excel_by_source(source: str):
    """Export work logs theo nguồn cụ thể (jira/bitbucket/git/manual)."""
    valid_sources = {"jira", "bitbucket", "github", "git", "manual"}
    if source not in valid_sources:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=400,
            detail=f"Invalid source. Must be one of: {', '.join(valid_sources)}",
        )
    return export_excel(source=source)
